#KEEP_LIST = ["PREFIX", "YRMO", "SEQ_NO", "PERMIT_NO", "APPLIED", "APPROVED", "ISSUED", "FINALED", "EXPIRED", "PARENT_PROJECT_NO", "PARENT_PERMIT_NO", "PermitType",
#             "PermitSubType", "STATUS", "SITE_LOT_NO", "SITE_BLOCK", "SITE_TRACT", "SITE_SUBDIVISION", "SITE_APN", "SITE_ADDR", "SITE_NUMBER", "SITE_STREETNAME",
#             "SITE_UNIT_NO", "SITE_CITY", "SITE_STATE", "SITE_ZIP", "DESCRIPTION", "OWNER_NAME", "APPLICANT_NAME", "CONTRACTOR_NAME", "SITE_ALTERNATE_ID",
#             "SITE_GEOTYPE"]
from openpyxl import load_workbook
import arcpy

loc = ("\\\\GISAPP\\Workspace\\GIS Staff Workspace\\cschultz\\NBRHD_DESC.xlsx") 
  
wb_obj = load_workbook(filename = loc)
wsheet = wb_obj['NBRHD_DESC']
dataDict = {}

for key, *values in wsheet.iter_rows():
    dataDict[key.value] = [v.value for v in values]

layer = "\\\\GISAPP\\Workspace\\GIS Staff Workspace\\cschultz\\2020PrelimParcels.gdb\\ParcelReport_1"
fields = ["hood_cd", "NBH"]
with arcpy.da.UpdateCursor(layer, fields) as Scur:
    for row in Scur:
        for key, value in dataDict.items():
            if row[0] == key:
                row[1] = value[0]
                Scur.updateRow(row)
print(".")